import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers

def generate_strict_outage_copy_report(file):
    xl = pd.ExcelFile(file)
    df = xl.parse(xl.sheet_names[0])

    df = df[~df['Status'].str.contains('fault cancel', case=False, na=False)]
    df = df[~df['Fault Type'].str.contains('fault cancelled', case=False, na=False)]
    df.fillna("unknown", inplace=True)
    df.replace(r'^\s*$', 'unknown', regex=True, inplace=True)

    if 'Outage' in df.columns:
        df.insert(df.columns.get_loc("Outage"), 'Hours', "")
    else:
        st.error("Expected column 'Outage' not found.")
        return None

    sheets = {
        'Valid': df[df['Fault Type'].str.lower() != '3rd party provider'],
        'Fiber': df[df['Fault Type'].str.lower().isin(['cable fault', 'other', 'others'])],
        'Power': df[df['Fault Type'].str.lower() == 'power problem'],
        'Equipment': df[df['Fault Type'].str.lower() == 'equipment fault'],
        '3rd Party': df[df['Fault Type'].str.lower() == '3rd party provider'],
    }

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for name in ['Valid', 'Fiber', 'Power', 'Equipment', '3rd Party']:
            sheet_df = sheets[name].copy()
            sheet_df.to_excel(writer, index=False, sheet_name=name)
            ws = writer.sheets[name]

            hours_col = sheet_df.columns.get_loc("Hours") + 1
            outage_col = sheet_df.columns.get_loc("Outage") + 1
            outage_letter = get_column_letter(outage_col)

            for row in range(2, len(sheet_df) + 2):
                outage_cell = ws.cell(row=row, column=outage_col)
                hours_cell = ws.cell(row=row, column=hours_col)

                if outage_cell.value == "unknown" or outage_cell.value in [None, "", "none"]:
                    outage_cell.value = "none"
                    continue

                if isinstance(outage_cell.value, str) and ":" in outage_cell.value:
                    outage_cell.number_format = "HH:MM"
                    hours_cell.value = f"={outage_letter}{row}*24"
                elif hasattr(outage_cell.value, 'hour'):
                    outage_cell.number_format = "HH:MM"
                    hours_cell.value = f"={outage_letter}{row}*24"

    output.seek(0)
    return output

# Streamlit UI
st.set_page_config(page_title="NDN Report Generator", layout="centered")
st.title("📊 NDN Report Generator")
st.write("Upload the raw Excel file to generate the cleaned report.")

uploaded_file = st.file_uploader("📂 Upload NDN Raw Excel File", type=["xlsx"])

if uploaded_file:
    with st.spinner("Processing..."):
        result = generate_strict_outage_copy_report(uploaded_file)
        if result:
            st.success("✅ Report generated successfully!")
            st.download_button(
                label="📥 Download Cleaned Report",
                data=result,
                file_name="NDN_2024_Cleaned_Report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
else:
    st.info("Please upload a valid Excel file to begin.")